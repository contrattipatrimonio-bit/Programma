import os
import sys
import json
import sqlite3
import pandas as pd
import shutil
import traceback
import secrets
import atexit
from functools import wraps
from datetime import datetime, timedelta
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, session, send_file, send_from_directory, make_response, Response
)
from werkzeug.utils import secure_filename
from dotenv import load_dotenv, set_key
from werkzeug.security import check_password_hash, generate_password_hash
from jinja2 import TemplateNotFound
import zipfile

# ==========================================================
# APERTURA IMMEDIATA DELLA PAGINA DI LOADING
# ==========================================================
import threading, time, webbrowser
def apri_loading_immediato():
    time.sleep(0.2)
    path_loading = os.path.abspath('loading.html')
    url = 'file://' + path_loading.replace('\\','/')
    try:
        webbrowser.open(url)
    except:
        pass
    print('🔵 Loading aperto subito:', url)

threading.Thread(target=apri_loading_immediato, daemon=True).start()

# ==========================================================
# 📁 Percorso base del programma (py o .exe su chiavetta/PC)
# ==========================================================
def get_base_dir():
    if getattr(sys, "frozen", False):
        # eseguibile generato (es. PyInstaller)
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_base_dir()

# ==========================================================
# 📡 CONFIGURAZIONE PORTATILE (offline + sync)
# ==========================================================

# Percorso cartella rete condivisa → da personalizzare
NETWORK_DATA_DIR = r"\\dp-smb\Supporto_Coordinamento\Compendio Atti"

# Cartella locale del programma (su PC o chiavetta USB)
LOCAL_DATA_DIR = os.path.join(BASE_DIR, "data_local")
os.makedirs(LOCAL_DATA_DIR, exist_ok=True)

LOCAL_DB = os.path.join(LOCAL_DATA_DIR, "compendio_norme.db")
LOCAL_EXCEL = os.path.join(LOCAL_DATA_DIR, "Elenconorme.xlsx")
LOCAL_PDF = os.path.join(LOCAL_DATA_DIR, "pdf")
os.makedirs(LOCAL_PDF, exist_ok=True)

def is_online():
    """Ritorna True se la cartella di rete è raggiungibile."""
    try:
        return os.path.exists(NETWORK_DATA_DIR)
    except:
        return False

def sync_from_network():
    """Copia da rete → locale SOLO se online."""
    if not is_online():
        print("⚠ Offline: uso database locale.")
        return False

    try:
        for fname in ["compendio_norme.db", "Elenconorme.xlsx"]:
            src = os.path.join(NETWORK_DATA_DIR, fname)
            dst = os.path.join(LOCAL_DATA_DIR, fname)
            if os.path.exists(src):
                shutil.copy2(src, dst)
                print("✔ Copiato da rete → locale:", fname)

        # Copia PDF
        net_pdf = os.path.join(NETWORK_DATA_DIR, "pdf")
        if os.path.exists(net_pdf):
            for f in os.listdir(net_pdf):
                if f.lower().endswith(".pdf"):
                    shutil.copy2(os.path.join(net_pdf, f),
                                 os.path.join(PDF_FOLDER, f))

        print("✔ Sincronizzazione rete → locale completata.")
        return True
    except Exception as e:
        print("❌ Errore sync rete → locale:", e)
        return False

def sync_to_network():
    """Copia locale → rete SOLO se online e hai il lock."""
    if not is_online():
        return False
    if not HAS_WRITE_LOCK:
        return False

    try:
        for fname in ["compendio_norme.db", "Elenconorme.xlsx"]:
            src = os.path.join(LOCAL_DATA_DIR, fname)
            dst = os.path.join(NETWORK_DATA_DIR, fname)
            if os.path.exists(src):
                shutil.copy2(src, dst)
                print("✔ Copiato locale → rete:", fname)

        # PDF
        net_pdf = os.path.join(NETWORK_DATA_DIR, "pdf")
        os.makedirs(net_pdf, exist_ok=True)
        for f in os.listdir(PDF_FOLDER):
            if f.lower().endswith(".pdf"):
                shutil.copy2(os.path.join(PDF_FOLDER, f),
                             os.path.join(net_pdf, f))

        print("✔ Sincronizzazione locale → rete completata.")
        return True
    except Exception as e:
        print("❌ Errore sync locale → rete:", e)
        return False

# ============================================================
#  LOCK PER SINGOLO ATTO (record-lock)
# ============================================================

LOCKS_DIR = os.path.join(NETWORK_DATA_DIR, "locks")

def get_record_lock_path(norma_id: int) -> str:
    return os.path.join(LOCKS_DIR, f"{norma_id}.lock")

def acquire_record_lock(norma_id: int) -> bool:
    """
    Tenta di acquisire un lock per l'atto <norma_id>.
    Se offline → nessun coordinamento possibile, si lascia lavorare (ritorna True).
    """
    if not is_online():
        return True

    os.makedirs(LOCKS_DIR, exist_ok=True)
    lock_path = get_record_lock_path(norma_id)

    # Controlla se lock già esistente è troppo vecchio
    if os.path.exists(lock_path):
        try:
            mtime = datetime.fromtimestamp(os.path.getmtime(lock_path))
            if datetime.now() - mtime > timedelta(hours=2):
                # Lock abbandonato da più di 2 ore → si elimina
                os.remove(lock_path)
            else:
                return False  # lock attivo da altro utente
        except Exception:
            return False

    # Crea nuovo lock
    try:
        info = (
            f"PC={os.environ.get('COMPUTERNAME','?')}; "
            f"USER={os.environ.get('USERNAME','?')}; "
            f"TS={datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        with open(lock_path, "w", encoding="utf-8") as f:
            f.write(info)
        return True
    except Exception:
        return False

def release_record_lock(norma_id: int):
    """Rilascia lock per l’atto <norma_id> se esiste."""
    if not is_online():
        return
    lock_path = get_record_lock_path(norma_id)
    try:
        if os.path.exists(lock_path):
            os.remove(lock_path)
    except Exception:
        pass
# ==========================================================
# 🔥 FUNZIONI PER LA GESTIONE DEI CONFLITTI DI SINCRONIZZAZIONE
# ==========================================================

CONFLICTS_FILE = os.path.join(LOCAL_DATA_DIR, "pending_conflicts.json")

def load_conflicts():
    """Carica la lista dei conflitti dal file JSON locale."""
    if not os.path.exists(CONFLICTS_FILE):
        return []
    try:
        with open(CONFLICTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return []

def save_conflicts(conflicts):
    """Salva i conflitti nel file JSON locale."""
    with open(CONFLICTS_FILE, "w", encoding="utf-8") as f:
        json.dump(conflicts, f, indent=2, ensure_ascii=False)

def add_conflict(key, local_row, network_row):
    """
    Aggiunge un nuovo conflitto nel file:
    key = "anno::numero::tipologia::fonte"
    """
    conflicts = load_conflicts()
    conflicts.append({
        "key": key,
        "local": local_row,
        "network": network_row
    })
    save_conflicts(conflicts)

# ==========================================================
# ⚙️ CONFIGURAZIONE INIZIALE
# ==========================================================
app = Flask(__name__)
app.secret_key = "chiave_segreta_sicura"  # 🔐 cambia in produzione

# 🔒 Token spegnimento via POST (route /__shutdown__)
SHUTDOWN_TOKEN = os.environ.get("SHUTDOWN_TOKEN") or secrets.token_hex(16)

# --- no-store per evitare cache
@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-store"
    return response

# ==========================================================
# 🖼️ favicon
# ==========================================================
@app.route('/favicon.ico')
def favicon():
    return send_from_directory(
        os.path.join(app.root_path, 'static'),
        'favicon.ico',
        mimetype='image/vnd.microsoft.icon'
    )

# ==========================================================
# 📁 utility percorso base (py/exe)
# ==========================================================

# 🔁 Cartella dati di rete (cartella condivisa)
# In questa versione portatile punta sempre alla cartella SMB configurata sopra.
DATA_DIR = NETWORK_DATA_DIR

# percorsi fondamentali:
# - il programma lavora SEMPRE sui file locali (LOCAL_*)
# - la cartella di rete viene usata solo per la sincronizzazione
DB_FILE = LOCAL_DB
EXCEL_FILE = LOCAL_EXCEL
PDF_FOLDER = os.path.join(BASE_DIR, "static", "pdf")
BACKUP_DIR = os.path.join(LOCAL_DATA_DIR, "backup")

# il log può restare locale a ogni PC
LOG_FILE = os.path.join(BASE_DIR, "app.log")

os.makedirs(PDF_FOLDER, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {"pdf"}

# ==========================================================
# 🔒 GESTIONE LOCK DI RETE (un solo "scrivente" alla volta)
# ==========================================================
LOCK_FILE = os.path.join(DATA_DIR, "compendio.lock")
HAS_WRITE_LOCK = False  # questa istanza può scrivere?

def acquire_lock():
    """Prova a creare il file di lock sulla cartella dati di rete."""
    global HAS_WRITE_LOCK
    if os.path.exists(LOCK_FILE):
        HAS_WRITE_LOCK = False
        return False
    try:
        info = (
            f"PC={os.environ.get('COMPUTERNAME', 'Sconosciuto')}; "
            f"USER={os.environ.get('USERNAME', 'Sconosciuto')}; "
            f"TS={datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        with open(LOCK_FILE, "w", encoding="utf-8") as f:
            f.write(info)
        HAS_WRITE_LOCK = True
        print(f"🔒 Lock acquisito: {info}")
        return True
    except Exception as e:
        print(f"⚠️ Impossibile creare il lock: {e}")
        HAS_WRITE_LOCK = False
        return False

def release_lock():
    """Rimuove il lock se questa istanza lo ha acquisito."""
    global HAS_WRITE_LOCK
    if HAS_WRITE_LOCK and os.path.exists(LOCK_FILE):
        try:
            os.remove(LOCK_FILE)
            print("🔓 Lock rilasciato.")
        except Exception as e:
            print(f"⚠️ Impossibile rimuovere il lock: {e}")
    HAS_WRITE_LOCK = False

# rilascio automatico del lock alla chiusura del processo
atexit.register(release_lock)

# ==========================================================
# 🛰️ sonda di avvio (per loading/launcher)
# ==========================================================
@app.get("/boot-ready.js")
def boot_ready_js():
    ok = True
    try:
        tpl = os.path.join(app.root_path, 'templates', 'menu_principale.html')
        if not os.path.exists(tpl):
            ok = False
        if not os.path.exists(DB_FILE):
            ok = False
        else:
            con = sqlite3.connect(DB_FILE)
            con.execute("SELECT 1")
            con.close()
    except Exception:
        ok = False

    js = f"window.__COMPENDIO_READY__ = {str(ok).lower()};"
    resp = make_response(js, 200)
    resp.mimetype = "application/javascript"
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return resp

# ==========================================================
# 🔧 .env e password amministratore
# ==========================================================
load_dotenv()
env_path = os.path.join(os.getcwd(), ".env")
if not os.path.exists(env_path) and os.path.exists(os.path.join(os.getcwd(), ".env.txt")):
    os.rename(os.path.join(os.getcwd(), ".env.txt"), env_path)
    print("✅ Rinomina automatica: .env.txt → .env")

dotenv_path = os.path.join(os.getcwd(), ".env")
if not os.path.exists(dotenv_path):
    print(f"⚠️ File .env non trovato in {dotenv_path}. Ne verrà creato uno nuovo.")
    open(dotenv_path, "a").close()

load_dotenv(dotenv_path, override=True)
raw_hash = os.getenv("ADMIN_PASSWORD_HASH")
if raw_hash:
    ADMIN_PASSWORD_HASH = raw_hash.strip().strip("'").strip('"')
    print(f"👉 DEBUG - Hash password letto da .env: {ADMIN_PASSWORD_HASH[:70]}...")
else:
    default_password = "admin1971"
    ADMIN_PASSWORD_HASH = generate_password_hash(default_password)
    set_key(dotenv_path, "ADMIN_PASSWORD_HASH", ADMIN_PASSWORD_HASH)
    print("⚠️ Nessuna password in .env — uso admin1971 di default.")
    print("✅ Creato nuovo hash per admin1971.")

# Seconda lettura .env (come da tuo script, lasciata)
dotenv_path2 = os.path.join(BASE_DIR, ".env")
if not os.path.exists(dotenv_path2):
    open(dotenv_path2, "a").close()
    print(f"⚠️ File .env non trovato — creato nuovo in {dotenv_path2}")

load_dotenv(dotenv_path2, override=True)
raw_hash2 = os.getenv("ADMIN_PASSWORD_HASH")
if raw_hash2:
    ADMIN_PASSWORD_HASH = raw_hash2.strip().strip("'").strip('"')
    print(f"👉 DEBUG - Hash password (ri)letto da .env: {ADMIN_PASSWORD_HASH[:70]}...")
else:
    default_password2 = "admin1971"
    ADMIN_PASSWORD_HASH = generate_password_hash(default_password2)
    set_key(dotenv_path2, "ADMIN_PASSWORD_HASH", ADMIN_PASSWORD_HASH)
    print("✅ Creato hash per password predefinita: admin1971 (secondo blocco)")

# ==========================================================
# 🔐 LOGIN
# ==========================================================
@app.route("/admin", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        password = request.form["password"]
        if ADMIN_PASSWORD_HASH and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session["admin"] = True
            log_event("admin_login", details={"result": "ok"})
            flash("✅ Accesso amministratore effettuato.", "success")
            return redirect(url_for("admin_dashboard"))
        else:
            log_event("admin_login_failed", details={"result": "bad_password"})
            flash("❌ Password errata.", "danger")
            return render_template("admin_login.html")
    return render_template("admin_login.html")

@app.route("/admin_alt", methods=["GET", "POST"])
def admin_login_alt():
    if request.method == "POST":
        password = request.form.get("password", "")
        if ADMIN_PASSWORD_HASH and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session["admin"] = True
            log_event("admin_login", details={"result": "ok", "route": "admin_alt"})
            flash("✅ Accesso amministratore effettuato.", "success")
            return redirect(url_for("admin_dashboard"))
        else:
            log_event("admin_login_failed", details={"result": "bad_password", "route": "admin_alt"})
            flash("❌ Password errata.", "danger")
            return render_template("admin_login.html")
    return render_template("admin_login.html")

# ==========================================================
# HOME
# ==========================================================
@app.route("/")
def home():
    return render_template("menu_principale.html")

# ==========================================================
# FUNZIONI DI SUPPORTO
# ==========================================================
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def backup_excel():
    if not os.path.exists(EXCEL_FILE):
        return
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    backup_name = f"Elenconorme_{timestamp}.xlsx"
    backup_path = os.path.join(BACKUP_DIR, backup_name)
    shutil.copy2(EXCEL_FILE, backup_path)
    print(f"✅ Backup Excel: {backup_path}")

def backup_pdf():
    if not os.path.exists(PDF_FOLDER):
        return
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    dest = os.path.join(BACKUP_DIR, f"Compendio_backup_{timestamp}")
    shutil.copytree(PDF_FOLDER, dest)
    print(f"✅ Backup PDF in {dest}")

def client_meta():
    # IP (anche dietro proxy)
    xff = request.headers.get("X-Forwarded-For", "")
    ip = (xff.split(",")[0].strip() if xff else request.remote_addr) or ""
    ua = request.headers.get("User-Agent", "")
    return ip, ua

def log_event(action, norma_id=None, details=None):
    """Scrive un evento in tabella audit."""
    try:
        ip, ua = client_meta()
        actor = "admin" if session.get("admin") else "anon"
        details_txt = json.dumps(details, ensure_ascii=False) if isinstance(details, (dict, list)) else (details or "")
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            INSERT INTO audit (action, norma_id, actor, ip, user_agent, details, ts)
            VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, (action, norma_id, actor, ip, ua, details_txt))
        conn.commit()
        conn.close()
    except Exception:
        traceback.print_exc()

def crea_database():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    # Tabella norme
    c.execute('''
        CREATE TABLE IF NOT EXISTS norme (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            anno TEXT,
            numero TEXT,
            tipologia TEXT,
            categoria TEXT,
            argomento TEXT,
            oggetto TEXT,
            fonte TEXT,
            filepdf TEXT,
            descrizione TEXT,
            stato TEXT,
            note TEXT
        )
    ''')
    # Tabella audit
    c.execute('''
        CREATE TABLE IF NOT EXISTS audit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT DEFAULT CURRENT_TIMESTAMP,
            action TEXT NOT NULL,
            norma_id INTEGER,
            actor TEXT,
            ip TEXT,
            user_agent TEXT,
            details TEXT
        )
    ''')
    conn.commit()
    conn.close()

def ensure_audit_table():
    """Chiamata aggiuntiva (idempotente) per garantire la tabella audit.
       È safe anche se già creata da crea_database()."""
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS audit (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT DEFAULT CURRENT_TIMESTAMP,
                action TEXT NOT NULL,
                norma_id INTEGER,
                actor TEXT,
                ip TEXT,
                user_agent TEXT,
                details TEXT
            )
        ''')
        conn.commit()
        conn.close()
    except Exception:
        traceback.print_exc()

def importa_dati_excel():
    if not os.path.exists(EXCEL_FILE):
        print("⚠️ File Excel non trovato.")
        return
    df = pd.read_excel(EXCEL_FILE)
    df.columns = [c.strip().lower() for c in df.columns]
    colonne_richieste = ["anno", "numero"]
    for col in colonne_richieste:
        if col not in df.columns:
            print(f"⚠️ Colonna mancante: {col}")
            return
    df = df[~(df["anno"].astype(str).str.strip().eq("") &
              df["numero"].astype(str).str.strip().eq(""))]
    df = df.fillna("").astype(str)
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("DELETE FROM norme")
    for _, row in df.iterrows():
        c.execute("""
            INSERT INTO norme (
                anno, numero, tipologia, categoria, argomento, oggetto,
                fonte, filepdf, descrizione, stato, note
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            row.get("anno", "").strip(),
            row.get("numero", "").strip(),
            row.get("tipologia", "").strip(),
            row.get("categoria", "").strip(),
            row.get("argomento", "").strip(),
            row.get("oggetto", "").strip(),
            row.get("fonte", "").strip(),
            row.get("filepdf", "").strip(),
            row.get("descrizione", "").strip(),
            row.get("stato", "").strip(),
            row.get("note", "").strip()
        ))
    conn.commit()
    conn.close()
    print(f"✅ Importazione completata: {len(df)} record importati da {EXCEL_FILE}.")

def require_write_lock(view_func):
    """Decorator: permette l'accesso in scrittura solo se questa istanza ha (o ottiene) il lock di rete.
    - Se siamo ONLINE e il lock non è disponibile → sola consultazione.
    - Se siamo OFFLINE → scrittura consentita su copia locale (senza lock di rete).
    """
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        global HAS_WRITE_LOCK
        # Se non abbiamo ancora il lock, decidiamo cosa fare ora
        if not HAS_WRITE_LOCK:
            if is_online():
                # ONLINE: proviamo ad acquisire il lock sulla cartella di rete
                if not acquire_lock():
                    flash("⚠️ Il sistema è in sola consultazione: un altro utente sta modificando i dati.", "warning")
                    return redirect(url_for("ricerca"))
            else:
                # OFFLINE: nessun lock di rete, ma permettiamo scrittura sulla copia locale
                HAS_WRITE_LOCK = True
        return view_func(*args, **kwargs)
    return wrapper

def admin_required(view_func):
    """Permette l'accesso solo se l'utente è admin (session['admin'] True)."""
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("admin"):
            flash("⚠️ Accesso non autorizzato.", "warning")
            return redirect(url_for("admin_login"))
        return view_func(*args, **kwargs)
    return wrapper


# ==========================================================
# 📊 DASHBOARD AMMINISTRATORE
# ==========================================================
@app.route("/admin_dashboard")
def admin_dashboard():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))

    # Conteggio atti
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM norme")
    totale_norme = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM norme WHERE filepdf IS NULL OR TRIM(filepdf) = ''")
    norme_senza_pdf = c.fetchone()[0]
    conn.close()

    # Conteggio PDF presenti in cartella
    pdf_dir = PDF_FOLDER
    totale_pdf = len([f for f in os.listdir(pdf_dir)
                      if f.lower().endswith(".pdf")]) if os.path.exists(pdf_dir) else 0

    # Data ultimo aggiornamento Excel
    if os.path.exists(EXCEL_FILE):
        data_ultimo_aggiornamento = datetime.fromtimestamp(
            os.path.getmtime(EXCEL_FILE)
        ).strftime("%d/%m/%Y %H:%M")
    else:
        data_ultimo_aggiornamento = "N/D"

    # Passo sia i nomi "nuovi" sia i nomi "vecchi" alla template
    return render_template(
        "admin_dashboard.html",
        totale_norme=totale_norme,
        norme_senza_pdf=norme_senza_pdf,
        totale_pdf=totale_pdf,
        data_ultimo_aggiornamento=data_ultimo_aggiornamento,
        # alias usati nella versione nuova dell'HTML
        totale=totale_norme,
        senza_pdf=norme_senza_pdf,
        ultimo_aggiornamento=data_ultimo_aggiornamento
    )
# ==========================================================
# 🩺 VERIFICA INTEGRITÀ DATABASE
# ==========================================================
@app.route("/verifica_integrita")
def verifica_integrita():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    problemi = []

    # Atti senza anno
    c.execute("SELECT COUNT(*) FROM norme WHERE anno IS NULL OR TRIM(anno) = ''")
    cnt_anno = c.fetchone()[0]
    if cnt_anno:
        problemi.append(f"{cnt_anno} atti senza anno.")

    # Atti senza numero
    c.execute("SELECT COUNT(*) FROM norme WHERE numero IS NULL OR TRIM(numero) = ''")
    cnt_num = c.fetchone()[0]
    if cnt_num:
        problemi.append(f"{cnt_num} atti senza numero.")

    # Atti senza oggetto
    c.execute("SELECT COUNT(*) FROM norme WHERE oggetto IS NULL OR TRIM(oggetto) = ''")
    cnt_ogg = c.fetchone()[0]
    if cnt_ogg:
        problemi.append(f"{cnt_ogg} atti senza oggetto.")

    # Duplicati anno/numero/fonte
    c.execute("""
        SELECT anno, numero, fonte, COUNT(*) as cnt
        FROM norme
        GROUP BY anno, numero, fonte
        HAVING cnt > 1
    """)
    duplicati = c.fetchall()
    if duplicati:
        problemi.append(f"{len(duplicati)} combinazioni anno/numero/fonte duplicate.")

    conn.close()

    if not problemi:
        flash("✅ Verifica completata: nessun problema di integrità evidente.", "success")
        log_event("verifica_integrita", details={"issues": []})
    else:
        msg = "⚠️ Verifica completata: " + " ".join(problemi)
        flash(msg, "warning")
        log_event("verifica_integrita", details={"issues": problemi})

    return redirect(url_for("admin_dashboard"))
# ==========================================================
# 📊 REPORT INCOERENZE PDF / DATABASE
# ==========================================================
@app.route("/report_incoerenze")
def report_incoerenze():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # Atti senza PDF associato
    c.execute("SELECT COUNT(*) FROM norme WHERE filepdf IS NULL OR TRIM(filepdf) = ''")
    cnt_senza_pdf = c.fetchone()[0]

    # Elenco PDF indicati nel DB
    c.execute("SELECT filepdf FROM norme WHERE filepdf IS NOT NULL AND TRIM(filepdf) <> ''")
    files_db_rows = c.fetchall()
    conn.close()

    files_db = {row[0] for row in files_db_rows if row[0]}
    pdf_mancanti = []
    pdf_orfani = []

    # PDF presenti sul disco ma non collegati
    if os.path.isdir(PDF_FOLDER):
        for f in os.listdir(PDF_FOLDER):
            if not f.lower().endswith(".pdf"):
                continue
            if f not in files_db:
                pdf_orfani.append(f)

    # PDF indicati nel DB ma non trovati in cartella
    for f in files_db:
        full_path = os.path.join(PDF_FOLDER, f)
        if not os.path.exists(full_path):
            pdf_mancanti.append(f)

    parti_msg = [f"{cnt_senza_pdf} atti senza PDF associato."]
    if pdf_mancanti:
        parti_msg.append(f"{len(pdf_mancanti)} PDF mancanti sul disco (presenti in DB ma non in cartella).")
    if pdf_orfani:
        parti_msg.append(f"{len(pdf_orfani)} PDF presenti in cartella ma non collegati ad alcun atto.")

    msg = "📊 Report incongruenze: " + " ".join(parti_msg)
    flash(msg, "info")

    # Log su audit (limito la lista per non esplodere)
    log_event("report_incoerenze_pdf", details={
        "atti_senza_pdf": cnt_senza_pdf,
        "pdf_mancanti": pdf_mancanti[:50],
        "pdf_orfani": pdf_orfani[:50]
    })

    return redirect(url_for("admin_dashboard"))
@app.route("/diagnostica_pdf")
@admin_required
def diagnostica_pdf():
    pdf_dir = PDF_FOLDER
    pdf_files = {f for f in os.listdir(pdf_dir) if f.lower().endswith(".pdf")}

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # Tutti gli atti con filepdf compilato
    atti = c.execute("SELECT id, anno, numero, filepdf FROM norme").fetchall()
    conn.close()

    # Atti senza PDF (campo DB vuoto)
    atti_senza_pdf = [a for a in atti if not a[3]]

    # PDF mancanti sul disco
    pdf_mancanti = [
        a for a in atti 
        if a[3] and a[3] not in pdf_files
    ]

    # PDF orfani (presenti sul disco ma non in DB)
    pdf_in_db = {a[3] for a in atti if a[3]}
    pdf_orfani = sorted(pdf_files - pdf_in_db)

    return render_template(
        "diagnostica_pdf.html",
        atti_senza_pdf=atti_senza_pdf,
        pdf_mancanti=pdf_mancanti,
        pdf_orfani=pdf_orfani
    )
@app.route("/diagnostica_pdf_export_csv")
@admin_required
def diagnostica_pdf_export_csv():
    import csv
    from io import StringIO

    pdf_dir = PDF_FOLDER
    pdf_files = {f for f in os.listdir(pdf_dir) if f.lower().endswith(".pdf")}

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    atti = c.execute("SELECT id, anno, numero, filepdf FROM norme").fetchall()
    conn.close()

    atti_senza_pdf = [a for a in atti if not a[3]]
    pdf_mancanti = [a for a in atti if a[3] and a[3] not in pdf_files]
    pdf_in_db = {a[3] for a in atti if a[3]}
    pdf_orfani = sorted(pdf_files - pdf_in_db)

    # Create CSV in memory
    output = StringIO()
    writer = csv.writer(output, delimiter=';')

    writer.writerow(["SEZIONE", "ID", "Anno", "Numero", "Nome PDF"])

    for a in atti_senza_pdf:
        writer.writerow(["ATTO SENZA PDF", a[0], a[1], a[2], ""])

    for a in pdf_mancanti:
        writer.writerow(["PDF MANCANTE", a[0], a[1], a[2], a[3]])

    for f in pdf_orfani:
        writer.writerow(["PDF ORFANO", "", "", "", f])

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            "Content-Disposition": "attachment; filename=diagnostica_pdf.csv"
        }
    )
@app.route("/diagnostica_pdf_export_xlsx")
@admin_required
def diagnostica_pdf_export_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from io import BytesIO

    pdf_dir = PDF_FOLDER
    pdf_files = {f for f in os.listdir(pdf_dir) if f.lower().endswith(".pdf")}

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    atti = c.execute("SELECT id, anno, numero, filepdf FROM norme").fetchall()
    conn.close()

    atti_senza_pdf = [a for a in atti if not a[3]]
    pdf_mancanti = [a for a in atti if a[3] and a[3] not in pdf_files]
    pdf_in_db = {a[3] for a in atti if a[3]}
    pdf_orfani = sorted(pdf_files - pdf_in_db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Diagnostica PDF"

    header = ["SEZIONE", "ID", "Anno", "Numero", "Nome PDF"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for a in atti_senza_pdf:
        ws.append(["ATTO SENZA PDF", a[0], a[1], a[2], ""])

    for a in pdf_mancanti:
        ws.append(["PDF MANCANTE", a[0], a[1], a[2], a[3]])

    for f in pdf_orfani:
        ws.append(["PDF ORFANO", "", "", "", f])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return Response(
        file_stream.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=diagnostica_pdf.xlsx"
        }
    )

# ==========================================================
# ⚠️ ADMIN — PAGINA DI GESTIONE CONFLITTI
# ==========================================================

@app.route("/admin/conflitti")
def admin_conflitti():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))

    conflitti = load_conflicts()
    return render_template("admin_conflitti.html", conflitti=conflitti)


@app.post("/admin/conflitti/risolvi/<int:conf_id>")
def risolvi_conflitto(conf_id):
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))

    conflitti = load_conflicts()
    if conf_id >= len(conflitti):
        flash("❌ Conflitto non valido.", "danger")
        return redirect(url_for("admin_conflitti"))

    conf = conflitti[conf_id]
    scelta = request.form.get("azione")

    # Estraggo chiavi
    anno, numero, tipologia, fonte = conf["key"].split("::")

    if scelta == "mantieni_locale":
        row = conf["local"]

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            UPDATE norme SET
                anno=?, numero=?, tipologia=?, argomento=?, oggetto=?,
                descrizione=?, stato=?, note=?, fonte=?, filepdf=?
            WHERE anno=? AND numero=? AND tipologia=? AND fonte=?
        """, (
            row["anno"], row["numero"], row["tipologia"], row["argomento"],
            row["oggetto"], row["descrizione"], row["stato"], row["note"],
            row["fonte"], row["filepdf"],
            anno, numero, tipologia, fonte
        ))
        conn.commit()
        conn.close()

        flash("✔ Versione LOCALE applicata.", "success")

    elif scelta == "mantieni_rete":
        row = conf["network"]

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            UPDATE norme SET
                anno=?, numero=?, tipologia=?, argomento=?, oggetto=?,
                descrizione=?, stato=?, note=?, fonte=?, filepdf=?
            WHERE anno=? AND numero=? AND tipologia=? AND fonte=?
        """, (
            row["anno"], row["numero"], row["tipologia"], row["argomento"],
            row["oggetto"], row["descrizione"], row["stato"], row["note"],
            row["fonte"], row["filepdf"],
            anno, numero, tipologia, fonte
        ))
        conn.commit()
        conn.close()

        flash("✔ Versione DI RETE ripristinata.", "success")

    elif scelta == "unisci":
        flash("🛠 Funzione di merge manuale in sviluppo.", "info")

    # Rimuovo il conflitto
    conflitti.pop(conf_id)
    save_conflicts(conflitti)

    return redirect(url_for("admin_conflitti"))

# ==========================================================
# ❌ ELIMINA UN SINGOLO ATTO + PDF
# ==========================================================
@app.route("/rimuovi_atto", methods=["GET", "POST"])
@require_write_lock
def rimuovi_atto():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))

    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # ------------------------------
    # POST: esegue l'eliminazione
    # ------------------------------
    if request.method == "POST":
        atto_id = request.form.get("atto_id", "").strip()
        if not atto_id:
            flash("❌ Nessun atto selezionato.", "error")
            conn.close()
            return redirect(url_for("rimuovi_atto"))

        # Recupero info atto (per log) + nome file PDF
        c.execute("SELECT id, anno, numero, fonte, oggetto, filepdf FROM norme WHERE id = ?", (atto_id,))
        row = c.fetchone()
        if not row:
            flash("❌ Atto non trovato.", "error")
            conn.close()
            return redirect(url_for("rimuovi_atto"))

        filepdf = row["filepdf"]
        info_atto = {
            "id": row["id"],
            "anno": row["anno"],
            "numero": row["numero"],
            "fonte": row["fonte"],
            "oggetto": row["oggetto"],
            "filepdf": filepdf,
        }

        # Elimino PDF se presente
        if filepdf:
            pdf_path = os.path.join(PDF_FOLDER, filepdf)
            if os.path.exists(pdf_path):
                try:
                    os.remove(pdf_path)
                except Exception:
                    traceback.print_exc()

        # Elimino riga dal database
        try:
            c.execute("DELETE FROM norme WHERE id = ?", (atto_id,))
            conn.commit()
            flash("✅ Atto eliminato correttamente (database + PDF).", "success")
        except Exception as e:
            traceback.print_exc()
            flash(f"❌ Errore durante l'eliminazione dell'atto: {e}", "error")
        finally:
            conn.close()

        # Log in audit
        log_event("delete_single_norma", norma_id=info_atto["id"], details=info_atto)

        return redirect(url_for("admin_dashboard"))

    # ------------------------------
    # GET: mostra form di ricerca
    # ------------------------------

    # Elenco fonti per il menù a tendina
    c.execute("""
        SELECT DISTINCT fonte
        FROM norme
        WHERE fonte IS NOT NULL AND TRIM(fonte) <> ''
        ORDER BY fonte
    """)
    fonti = [row["fonte"] for row in c.fetchall()]

    # Se sono stati passati criteri di ricerca, cerco gli atti corrispondenti
    anno = request.args.get("anno", "").strip()
    numero = request.args.get("numero", "").strip()
    fonte_sel = request.args.get("fonte", "").strip()

    risultati = []
    if anno or numero or fonte_sel:
        query = "SELECT id, anno, numero, fonte, oggetto FROM norme WHERE 1=1"
        params = []
        if anno:
            query += " AND anno = ?"
            params.append(anno)
        if numero:
            query += " AND numero = ?"
            params.append(numero)
        if fonte_sel:
            query += " AND fonte = ?"
            params.append(fonte_sel)

        query += " ORDER BY anno DESC, numero ASC"
        c.execute(query, params)
        risultati = c.fetchall()

    conn.close()

    return render_template(
        "admin_rimuovi_atto.html",
        fonti=fonti,
        risultati=risultati
    )

# ==========================================================
# 🔑 CAMBIO PASSWORD
# ==========================================================
@app.route("/cambia_password", methods=["GET", "POST"])
def cambia_password():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))

    global ADMIN_PASSWORD_HASH
    dotenv_path_local = os.path.join(os.getcwd(), ".env")

    if request.method == "POST":
        old_password = request.form.get("old_password", "").strip()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        if not check_password_hash(ADMIN_PASSWORD_HASH, old_password):
            flash("❌ Password attuale errata.", "danger")
            return redirect(url_for("cambia_password"))

        if new_password != confirm_password:
            flash("⚠️ Le nuove password non coincidono.", "warning")
            return redirect(url_for("cambia_password"))

        new_hash = generate_password_hash(new_password)
        set_key(dotenv_path_local, "ADMIN_PASSWORD_HASH", new_hash)
        ADMIN_PASSWORD_HASH = new_hash
        flash("✅ Password aggiornata con successo!", "success")
        return redirect(url_for("admin_dashboard"))

    return render_template("cambia_password.html")

# ==========================================================
# LOG / IMPORT / CLEAR / BACKUP / ZIP PDF
# ==========================================================
@app.route("/download_log")
def download_log():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))
    if not os.path.exists(LOG_FILE):
        flash("⚠️ Nessun file di log presente.", "warning")
        return redirect(url_for("admin_dashboard"))
    return send_file(LOG_FILE, as_attachment=True, download_name="app_log.txt")

@app.route("/clear_log")
@require_write_lock
def clear_log():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))
    open(LOG_FILE, "w").close()
    flash("🧹 Log svuotato correttamente.", "success")
    return redirect(url_for("admin_dashboard"))

@app.route("/import_excel")
@require_write_lock
def import_excel():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))
    if not os.path.exists(EXCEL_FILE):
        flash("❌ File Excel non trovato.", "error")
        return redirect(url_for("admin_dashboard"))
    try:
        df = pd.read_excel(EXCEL_FILE)
        df.columns = [c.strip().lower() for c in df.columns]
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        nuovi, aggiornati = 0, 0
        for _, row in df.iterrows():
            anno = str(row.get("anno", "")).strip()
            numero = str(row.get("numero", "")).strip()
            if not anno or not numero:
                continue
            esiste = c.execute("SELECT COUNT(*) FROM norme WHERE anno=? AND numero=?", (anno, numero)).fetchone()[0]
            if esiste:
                aggiornati += 1
                continue
            else:
                nuovi += 1
                c.execute("""
                    INSERT INTO norme (anno, numero, tipologia, categoria, argomento, oggetto,
                                       fonte, filepdf, descrizione, stato, note)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    anno,
                    numero,
                    str(row.get("tipologia", "")),
                    str(row.get("categoria", "")),
                    str(row.get("argomento", "")),
                    str(row.get("oggetto", "")),
                    str(row.get("fonte", "")),
                    str(row.get("filepdf", "")),
                    str(row.get("descrizione", "")),
                    str(row.get("stato", "")),
                    str(row.get("note", "")),
                ))
        conn.commit()
        conn.close()
        flash(f"✅ Importazione completata: {nuovi} nuovi atti aggiunti, {aggiornati} già presenti.", "success")
    except Exception as e:
        traceback.print_exc()
        flash(f"❌ Errore durante l’importazione Excel: {e}", "error")
    return redirect(url_for("admin_dashboard"))

@app.route("/upload_excel", methods=["GET", "POST"])
@require_write_lock
def upload_excel():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))

    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename.lower().endswith(".xlsx"):
            flash("❌ Caricare un file Excel valido (.xlsx).", "error")
            return redirect(url_for("upload_excel"))

        try:
            upload_path = os.path.join(BASE_DIR, "ElencoNorme_uploaded.xlsx")
            file.save(upload_path)
            df = pd.read_excel(upload_path)
            df.columns = [c.strip().lower() for c in df.columns]
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            nuovi, aggiornati = 0, 0
            for _, row in df.iterrows():
                anno = str(row.get("anno", "")).strip()
                numero = str(row.get("numero", "")).strip()
                if not anno or not numero:
                    continue
                esiste = c.execute("SELECT COUNT(*) FROM norme WHERE anno=? AND numero=?", (anno, numero)).fetchone()[0]
                if esiste:
                    aggiornati += 1
                else:
                    nuovi += 1
                    c.execute("""
                        INSERT INTO norme (anno, numero, tipologia, categoria, argomento, oggetto,
                                           fonte, filepdf, descrizione, stato, note)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        anno,
                        numero,
                        str(row.get("tipologia", "")),
                        str(row.get("categoria", "")),
                        str(row.get("argomento", "")),
                        str(row.get("oggetto", "")),
                        str(row.get("fonte", "")),
                        str(row.get("filepdf", "")),
                        str(row.get("descrizione", "")),
                        str(row.get("stato", "")),
                        str(row.get("note", "")),
                    ))
            conn.commit()
            conn.close()
            os.remove(upload_path)
            flash(f"✅ Importazione completata: {nuovi} nuovi atti, {aggiornati} già presenti.", "success")
        except Exception as e:
            traceback.print_exc()
            flash(f"❌ Errore durante l’importazione Excel: {e}", "error")
        return redirect(url_for("admin_dashboard"))

    return """
    <h2 style='font-family:sans-serif;text-align:center;margin-top:40px;'>📤 Carica file Excel</h2>
    <form method='POST' enctype='multipart/form-data' style='text-align:center;margin-top:30px;'>
        <input type='file' name='file' accept='.xlsx' required>
        <br><br>
        <button type='submit' style='padding:10px 20px;background:#0073e6;color:white;border:none;border-radius:6px;cursor:pointer;'>📊 Importa Excel</button>
    </form>
    <p style='text-align:center;margin-top:20px;'>
        <a href='/admin_dashboard'>⬅️ Torna alla Dashboard</a>
    </p>
    """

@app.route("/clear_database")
@require_write_lock
def clear_database():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("DELETE FROM norme")
        conn.commit()
        conn.close()
        for f in os.listdir(PDF_FOLDER):
            if f.lower().endswith(".pdf"):
                os.remove(os.path.join(PDF_FOLDER, f))
        flash("💣 Database e PDF svuotati correttamente.", "success")
    except Exception as e:
        traceback.print_exc()
        flash(f"❌ Errore durante la pulizia del database: {e}", "error")
    return redirect(url_for("admin_dashboard"))

@app.route("/fix_nan_db")
@require_write_lock
def fix_nan_db():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            UPDATE norme SET
                descrizione = CASE WHEN lower(descrizione) = 'nan' THEN '' ELSE descrizione END,
                note        = CASE WHEN lower(note) = 'nan' THEN '' ELSE note END,
                oggetto     = CASE WHEN lower(oggetto) = 'nan' THEN '' ELSE oggetto END,
                argomento   = CASE WHEN lower(argomento) = 'nan' THEN '' ELSE argomento END,
                tipologia   = CASE WHEN lower(tipologia) = 'nan' THEN '' ELSE tipologia END,
                fonte       = CASE WHEN lower(fonte) = 'nan' THEN '' ELSE fonte END,
                filepdf     = CASE WHEN lower(filepdf) = 'nan' THEN '' ELSE filepdf END
        """)
        conn.commit()
        conn.close()
        flash("🧹 Database ripulito da tutti i valori 'nan'.", "success")
    except Exception as e:
        traceback.print_exc()
        flash(f"❌ Errore durante la pulizia DB: {e}", "danger")
    return redirect(url_for("admin_dashboard"))

@app.route("/backup_pdfs")
@require_write_lock
def backup_pdfs():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        zip_name = f"backup_pdf_{timestamp}.zip"
        zip_path = os.path.join(BACKUP_DIR, zip_name)
        os.makedirs(BACKUP_DIR, exist_ok=True)
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(PDF_FOLDER):
                for file in files:
                    if file.lower().endswith(".pdf"):
                        full_path = os.path.join(root, file)
                        arcname = os.path.relpath(full_path, PDF_FOLDER)
                        zipf.write(full_path, arcname)
        flash(f"🗃️ Backup PDF creato: {zip_name}", "success")
        return send_file(zip_path, as_attachment=True)
    except Exception as e:
        traceback.print_exc()
        flash(f"❌ Errore durante il backup PDF: {e}", "error")
        return redirect(url_for("admin_dashboard"))

@app.route("/import_pdfs", methods=["GET", "POST"])
@require_write_lock
def import_pdfs():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))

    if request.method == "POST":
        file = request.files.get("zipfile")
        if not file or not file.filename.lower().endswith(".zip"):
            flash("❌ Caricare un file ZIP valido.", "error")
            return redirect(url_for("import_pdfs"))
        try:
            temp_path = os.path.join(BASE_DIR, "temp_import.zip")
            file.save(temp_path)
            with zipfile.ZipFile(temp_path, "r") as zip_ref:
                zip_ref.extractall(PDF_FOLDER)
            os.remove(temp_path)
            flash("✅ PDF importati con successo nella cartella PDF condivisa.", "success")
        except Exception as e:
            traceback.print_exc()
            flash(f"❌ Errore durante l’importazione dei PDF: {e}", "error")
        return redirect(url_for("admin_dashboard"))

    return """
    <h2 style='font-family:sans-serif;text-align:center;margin-top:40px;'>📦 Importa PDF da ZIP</h2>
    <form method="POST" enctype="multipart/form-data" style="text-align:center;margin-top:30px;">
        <input type="file" name="zipfile" accept=".zip" required>
        <br><br>
        <button type="submit" style="padding:10px 20px;background:#b30000;color:white;border:none;border-radius:6px;cursor:pointer;">📤 Carica ZIP</button>
    </form>
    <p style="text-align:center;margin-top:20px;"><a href='/admin_dashboard'>⬅️ Torna alla dashboard</a></p>
    """

# ==========================================================
# 🔒 LOGOUT – chiusura immediata, senza redirect o template
# ==========================================================
@app.get("/logout")
def logout():
    """Logout con pagina di uscita grafica."""
    log_event("logout", details={"where": "menu"})
    session.pop("admin", None)
    return render_template("uscita.html")

    html = """
    <!DOCTYPE html>
    <html lang="it">
    <head>
      <meta charset="utf-8">
      <title>Chiusura</title>
      <script>
        // Chiude la finestra dopo 100 ms
        setTimeout(() => { 
          window.open('', '_self', ''); 
          window.close(); 
        }, 100);

        // Fallback se il browser blocca window.close()
        setTimeout(() => {
          document.body.innerHTML = `
            <div style="font-family:sans-serif;text-align:center;margin-top:60px;">
              <h2>🔒 Sessione terminata</h2>
              <p>Puoi ora chiudere questa scheda manualmente.</p>
            </div>`;
        }, 600);
      </script>
    </head>
    <body style="background:#fafafa;"></body>
    </html>
    """
    return html

# ==========================================================
# 🛑 SPEGNIMENTO SERVER (opzionale)
# ==========================================================
@app.post("/__shutdown__")
def shutdown():
    """Permette di chiudere il server Flask in modo controllato."""
    token = request.headers.get("X-Shutdown-Token") or request.form.get("token")
    if token != SHUTDOWN_TOKEN:
        return "Forbidden", 403

    func = request.environ.get("werkzeug.server.shutdown")
    if not func:
        return "Not running with the Werkzeug Server", 501

    print("🛑 Arresto del server richiesto…")
    func()
    return "OK", 200

# ==========================================================
# 🔍 RICERCA / ESPORTAZIONE
# ==========================================================
@app.route("/ricerca")
def ricerca():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    tipologie = [r["tipologia"] for r in c.execute(
        "SELECT DISTINCT tipologia FROM norme WHERE tipologia != '' ORDER BY tipologia").fetchall()]
    argomenti = [r["argomento"] for r in c.execute(
        "SELECT DISTINCT argomento FROM norme WHERE argomento != '' ORDER BY argomento").fetchall()]
    fonti = [r["fonte"] for r in c.execute(
        "SELECT DISTINCT fonte FROM norme WHERE fonte != '' ORDER BY fonte").fetchall()]

    query = "SELECT * FROM norme WHERE 1=1"
    params = []

    # Filtri base
    if request.args.get("anno"):
        query += " AND anno LIKE ?"
        params.append(f"%{request.args.get('anno')}%")
    if request.args.get("numero"):
        query += " AND numero = ?"
        params.append(request.args.get("numero"))
    if request.args.get("tipo") and request.args.get("tipo") != "Tutto":
        query += " AND tipologia=?"
        params.append(request.args.get("tipo"))
    if request.args.get("argomento") and request.args.get("argomento") != "Tutto":
        query += " AND argomento LIKE ?"
        params.append(f"%{request.args.get('argomento')}%")
    if request.args.get("fonte") and request.args.get("fonte") != "Tutto":
        query += " AND fonte=?"
        params.append(request.args.get("fonte"))

    # Ricerca "Google-like" sul testo
    testo = (request.args.get("testo") or "").strip()
    if testo:
        tokens = testo.split()
        testo_clauses = []
        testo_params = []

        # colonne su cui fare ricerca testuale
        searchable_cols = ["oggetto", "descrizione", "argomento", "tipologia", "fonte", "note"]

        for token in tokens:
            if token.startswith("-") and len(token) > 1:
                # parola da escludere
                word = token[1:]
                sub_parts = []
                for col in searchable_cols:
                    sub_parts.append(f"{col} NOT LIKE ?")
                    testo_params.append(f"%{word}%")
                testo_clauses.append("(" + " AND ".join(sub_parts) + ")")
            else:
                # parola (o frase) da cercare
                word = token
                sub_parts = []
                for col in searchable_cols:
                    sub_parts.append(f"{col} LIKE ?")
                    testo_params.append(f"%{word}%")
                testo_clauses.append("(" + " OR ".join(sub_parts) + ")")

        if testo_clauses:
            query += " AND (" + " AND ".join(testo_clauses) + ")"
            params.extend(testo_params)

    norme = c.execute(query + " ORDER BY anno DESC, numero ASC", params).fetchall()

    # conteggio totale se servisse
    c.execute("SELECT COUNT(*) FROM norme")
    totale = c.fetchone()[0]

    conn.close()
    return render_template(
        "ricerca.html",
        norme=norme,
        tipologie=tipologie,
        categorie=argomenti,
        fonti=fonti,
        totale=totale
    )

@app.route("/esportazione")
def esportazione():
    tipo = request.args.get("tipo", "risultati")
    filtri = request.args.to_dict()
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    query = "SELECT COUNT(*) as totale FROM norme WHERE 1=1"
    params = []
    if tipo == "risultati":
        if filtri.get("anno"):
            query += " AND anno LIKE ?"
            params.append(f"%{filtri['anno']}%")
        if filtri.get("numero"):
            query += " AND numero LIKE ?"
            params.append(filtri["numero"])
        if filtri.get("tipologia") and filtri["tipologia"] != "Tutto":
            query += " AND tipologia = ?"
            params.append(filtri["tipologia"])
        if filtri.get("argomento") and filtri["argomento"] != "Tutto":
            query += " AND argomento LIKE ?"
            params.append(f"%{filtri['argomento']}%")
        if filtri.get("fonte") and filtri["fonte"] != "Tutto":
            query += " AND fonte = ?"
            params.append(filtri["fonte"])
        if filtri.get("testo"):
            query += " AND (oggetto LIKE ? OR descrizione LIKE ?)"
            params.extend([f"%{filtri['testo']}%", f"%{filtri['testo']}%"])
    totale = c.execute(query, params).fetchone()["totale"]
    conn.close()
    return render_template("esportazione.html", tipo=tipo, totale=totale, filtri=filtri)

# ==========================================================
# 🧾 ESECUZIONE DELL'ESPORTAZIONE (Excel / PDF)
# ==========================================================
@app.route("/esegui_export", methods=["POST"])
def esegui_export():
    import io
    from flask import send_file as send_file_flask
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    tipo = request.form.get("tipo", "risultati")
    campi = request.form.getlist("campi")
    formato = request.form.get("formato", "excel")
    filtri = {k: v for k, v in request.form.items() if k not in ["campi", "formato"] and k != "tipo"}

    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    query = "SELECT * FROM norme WHERE 1=1"
    params = []
    if tipo == "risultati":
        if filtri.get("anno"):
            query += " AND CAST(anno AS TEXT) LIKE ?"
            params.append(f"%{filtri['anno']}%")
        if filtri.get("numero"):
            query += " AND CAST(numero AS TEXT) LIKE ?"
            params.append(f"{filtri['numero']}")
        if filtri.get("tipologia") and filtri["tipologia"] != "Tutto":
            query += " AND tipologia = ?"
            params.append(filtri["tipologia"])
        if filtri.get("argomento") and filtri["argomento"] != "Tutto":
            query += " AND argomento LIKE ?"
            params.append(f"%{filtri['argomento']}%")
        if filtri.get("fonte") and filtri["fonte"] != "Tutto":
            query += " AND fonte = ?"
            params.append(filtri["fonte"])
        if filtri.get("testo"):
            query += " AND (oggetto LIKE ? OR descrizione LIKE ?)"
            params.extend([f"%{filtri['testo']}%", f"%{filtri['testo']}%"])
    dati = c.execute(query + " ORDER BY anno DESC, numero ASC", params).fetchall()
    conn.close()

    if not dati:
        flash("⚠️ Nessun dato trovato per l'esportazione.", "warning")
        return redirect(url_for("ricerca"))

    if not campi:
        campi = ["Anno", "Numero", "Tipologia", "Oggetto"]

    if formato == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Atti Esportati"
        rosso_milano = "B30000"
        bianco = "FFFFFF"
        grigio_chiaro = "F5F5F5"
        bold_font = Font(bold=True, color=bianco)
        header_fill = PatternFill(start_color=rosso_milano, end_color=rosso_milano, fill_type="solid")
        normal_fill = PatternFill(start_color=bianco, end_color=bianco, fill_type="solid")
        alt_fill = PatternFill(start_color=grigio_chiaro, end_color=grigio_chiaro, fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )
        last_col_letter = get_column_letter(len(campi))
        ws.merge_cells(f"A1:{last_col_letter}1")
        cell_header = ws["A1"]
        cell_header.value = "COMUNE DI MILANO – Sistema GESTIONE ATTI"
        cell_header.font = Font(bold=True, color="FFFFFF", size=14)
        cell_header.fill = header_fill
        cell_header.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        ws.append(campi)
        for cell in ws[2]:
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for i, row in enumerate(dati, start=3):
            for j, campo in enumerate(campi, start=1):
                key = campo.strip().lower()
                val = str(row[key]) if key in row.keys() and row[key] is not None else ""
                cell = ws.cell(row=i, column=j, value=val)
                cell.fill = alt_fill if (i % 2 == 0) else normal_fill
                cell.border = thin_border
                if key in ["anno", "numero", "stato"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        col_widths = {}
        for idx, campo in enumerate(campi, start=1):
            nome = campo.lower()
            if nome in ["anno"]:
                col_widths[idx] = 10
            elif nome in ["numero"]:
                col_widths[idx] = 12
            elif nome in ["tipologia", "fonte", "argomento"]:
                col_widths[idx] = 20
            elif nome in ["oggetto"]:
                col_widths[idx] = 40
            elif nome in ["descrizione"]:
                col_widths[idx] = 45
            elif nome in ["stato"]:
                col_widths[idx] = 15
            elif nome in ["note"]:
                col_widths[idx] = 25
            else:
                col_widths[idx] = 18
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        binary_excel = io.BytesIO()
        wb.save(binary_excel)
        binary_excel.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file_flask(
            binary_excel,
            as_attachment=True,
            download_name=f"atti_{tipo}_{timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    elif formato == "pdf":
        import io
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        binary_pdf = io.BytesIO()
        page_width, page_height = landscape(A4)
        doc = SimpleDocTemplate(
            binary_pdf,
            pagesize=(page_width, page_height),
            leftMargin=1 * cm,
            rightMargin=1 * cm,
            topMargin=2.5 * cm,
            bottomMargin=2 * cm,
        )
        styles = getSampleStyleSheet()
        story = []

        from reportlab.lib import colors

        def draw_header_footer(canvas, doc_obj):
            canvas.setFillColor(colors.HexColor("#B30000"))
            canvas.rect(0, page_height - 50, page_width, 25, stroke=0, fill=1)
            canvas.setFillColor(colors.white)
            canvas.setFont("Helvetica-Bold", 13)
            canvas.drawString(2 * cm, page_height - 42, "COMUNE DI MILANO – Sistema GESTIONE ATTI")
            canvas.setFont("Helvetica", 9)
            canvas.setFillColor(colors.gray)
            canvas.drawString(2 * cm, 1 * cm, "Comune di Milano – Sistema GESTIONE ATTI")
            canvas.drawRightString(page_width - 2 * cm, 1 * cm, datetime.now().strftime("%d/%m/%Y %H:%M"))

        story.append(Spacer(1, 1 * cm))
        story.append(Paragraph("<b>Archivio Atti Esportato</b>", styles["Heading2"]))
        story.append(Paragraph("Esportazione generata automaticamente dal sistema <b>GESTIONE ATTI</b>.", styles["Normal"]))
        story.append(Spacer(1, 0.5 * cm))

        if dati:
            from reportlab.lib.enums import TA_LEFT
            normal_style = styles["Normal"]
            normal_style.fontSize = 8
            normal_style.leading = 10
            normal_style.alignment = TA_LEFT
            header_style = ParagraphStyle(
                "HeaderStyle",
                parent=styles["Normal"],
                fontName="Helvetica-Bold",
                fontSize=9,
                textColor=colors.white,
                alignment=1
            )
            data = [[Paragraph(campo, header_style) for campo in campi]]
            for row in dati:
                riga = []
                for c_field in campi:
                    key = c_field.lower()
                    val = ""
                    if key in row.keys() and row[key]:
                        val = str(row[key])
                    riga.append(Paragraph(val.replace("\n", "<br/>"), normal_style))
                data.append(riga)

            usable_width = page_width - 2 * cm
            col_widths_guess = []
            for name in campi:
                nome = name.lower()
                if nome in ["anno"]:
                    col_widths_guess.append(2.0 * cm)
                elif nome in ["numero"]:
                    col_widths_guess.append(2.5 * cm)
                elif nome in ["tipologia", "fonte", "argomento"]:
                    col_widths_guess.append(4.0 * cm)
                elif nome in ["stato"]:
                    col_widths_guess.append(2.5 * cm)
                elif nome in ["note"]:
                    col_widths_guess.append(5.0 * cm)
                elif nome in ["oggetto"]:
                    col_widths_guess.append(7.0 * cm)
                elif nome in ["descrizione"]:
                    col_widths_guess.append(8.0 * cm)
                else:
                    col_widths_guess.append(usable_width / len(campi))
            scale = usable_width / sum(col_widths_guess)
            col_widths = [w * scale for w in col_widths_guess]

            table = Table(data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B30000")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                ("VALIGN", (0, 1), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.Color(0.94, 0.94, 0.94)]),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.Color(0.8, 0.8, 0.8)),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(Spacer(1, 6))
            story.append(table)
        else:
            story.append(Paragraph("Nessun atto disponibile per l’esportazione.", styles["Normal"]))

        doc.build(story, onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)
        binary_pdf.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file_flask(
            binary_pdf,
            as_attachment=True,
            download_name=f"atti_{tipo}_{timestamp}.pdf",
            mimetype="application/pdf"
        )

    flash("Formato esportazione non valido.", "error")
    return redirect(url_for("ricerca"))
# ==========================================================
# 📊 DETTAGLIO / MODIFICA / INSERIMENTO ATTI (con Audit)
# ==========================================================
def registra_audit(azione, norma_id, old_row, new_data):
    """Registra un evento di modifica in audit, salvando differenze."""
    try:
        diff = {}
        if old_row:
            # vecchia riga sqlite3.Row → dict
            old_dict = {}
            for k in old_row.keys():
                old_dict[k] = old_row[k]
            for k, new_val in new_data.items():
                old_val = old_dict.get(k)
                if (old_val or "") != (new_val or ""):
                    diff[k] = {"old": old_val, "new": new_val}
        else:
            diff = {"new": new_data}
        log_event(azione, norma_id=norma_id, details=diff)
    except Exception:
        traceback.print_exc()

def aggiorna_excel_singolo(norma_id):
    """Aggiorna (se esiste) la riga corrispondente in EXCEL_FILE per l'atto indicato."""
    if not os.path.exists(EXCEL_FILE):
        return
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        r = c.execute("SELECT * FROM norme WHERE id=?", (norma_id,)).fetchone()
        conn.close()
        if not r:
            return
        df = pd.read_excel(EXCEL_FILE)
        # normalizza nomi
        cols = [c.strip().lower() for c in df.columns]
        colmap = {c.lower(): i for i, c in enumerate(df.columns)}
        # match per anno + numero + fonte
        mask = (
            df.iloc[:, colmap.get("anno")].astype(str).str.strip() == str(r["anno"]).strip()
        ) & (
            df.iloc[:, colmap.get("numero")].astype(str).str.strip() == str(r["numero"]).strip()
        )
        if "fonte" in colmap:
            mask &= df.iloc[:, colmap["fonte"]].astype(str).str.strip() == str(r["fonte"]).strip()
        if not mask.any():
            # non trovato → opzionale: append
            return
        idx = df[mask].index
        def set_if_col(colname, value):
            ci = colmap.get(colname)
            if ci is not None:
                df.iloc[idx, ci] = value
        set_if_col("anno", r["anno"])
        set_if_col("numero", r["numero"])
        set_if_col("tipologia", r["tipologia"])
        set_if_col("argomento", r["argomento"])
        set_if_col("oggetto", r["oggetto"])
        set_if_col("descrizione", r["descrizione"])
        set_if_col("stato", r["stato"])
        set_if_col("note", r["note"])
        set_if_col("fonte", r["fonte"])
        set_if_col("filepdf", r["filepdf"])
        df.to_excel(EXCEL_FILE, index=False)
    except Exception:
        traceback.print_exc()

@app.route("/dettaglio/<int:norma_id>")
def dettaglio(norma_id):
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    norma = c.execute("SELECT * FROM norme WHERE id=?", (norma_id,)).fetchone()
    conn.close()

    # Pulizia NAN + None
    if norma:
        norma = dict(norma)
        for k, v in norma.items():
            if v is None or str(v).strip().lower() == "nan":
                norma[k] = ""

    return render_template("dettaglio.html", norma=norma)

@app.route("/modifica/<int:norma_id>", methods=["GET", "POST"])
@require_write_lock
def modifica(norma_id):
    """Modifica atto con blocco per singolo record."""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # -------------------------
    # SALVATAGGIO (POST)
    # -------------------------
    if request.method == "POST":
        # Dati vecchi
        old = c.execute("SELECT * FROM norme WHERE id=?", (norma_id,)).fetchone()

        # Nuovi dati
        data = {k: request.form.get(k, "").strip() for k in
                ["anno", "numero", "tipologia", "argomento", "oggetto",
                 "descrizione", "stato", "note", "fonte"]}

        # UPDATE
        c.execute("""
            UPDATE norme SET
                anno=?, numero=?, tipologia=?, argomento=?, oggetto=?, 
                descrizione=?, stato=?, note=?, fonte=?
            WHERE id=?
        """, (
            data["anno"], data["numero"], data["tipologia"], data["argomento"],
            data["oggetto"], data["descrizione"], data["stato"], data["note"],
            data["fonte"], norma_id
        ))
        conn.commit()

        # Aggiornamento Excel locale
        aggiorna_excel_singolo(norma_id)

        # Audit
        registra_audit("modifica", norma_id, old, data)

        conn.close()

        # 🔓 Rilascio lock per QUESTO ATTO
        release_record_lock(norma_id)

        # 🔓 Rilascio lock globale
        release_lock()

        flash("✅ Modifica salvata con successo!", "success")
        return redirect(url_for("dettaglio", norma_id=norma_id))

    # -------------------------
    # APERTURA FORM (GET)
    # -------------------------

    # Prova a prendere il lock per questo atto
    if not acquire_record_lock(norma_id):
        conn.close()
        flash("⚠️ Questo atto è attualmente in modifica da un altro utente.", "warning")
        return redirect(url_for("dettaglio", norma_id=norma_id))

    # Carica record e mostra pagina di modifica
    norma = c.execute("SELECT * FROM norme WHERE id=?", (norma_id,)).fetchone()
    conn.close()

    return render_template("modifica.html", norma=norma)

@app.route("/inserisci", methods=["GET", "POST"])
@require_write_lock
def inserisci():
    # combobox valori
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    tipologie = [r["tipologia"] for r in c.execute(
        "SELECT DISTINCT tipologia FROM norme WHERE tipologia != '' ORDER BY tipologia").fetchall()]
    argomenti = [r["argomento"] for r in c.execute(
        "SELECT DISTINCT argomento FROM norme WHERE argomento != '' ORDER BY argomento").fetchall()]
    fonti = [r["fonte"] for r in c.execute(
        "SELECT DISTINCT fonte FROM norme WHERE fonte != '' ORDER BY fonte").fetchall()]
    conn.close()

    if request.method == "POST":
        campi = ["anno", "numero", "tipologia", "argomento", "oggetto",
                 "descrizione", "stato", "note", "fonte"]
        dati = []
        for campo in campi:
            valore = request.form.get(campo, "").strip()
            if valore.lower() == "altro" or valore == "":
                valore_altro = request.form.get(f"altro_{campo}", "").strip()
                if valore_altro:
                    valore = valore_altro
            if valore == "":
                valore = "-"
            dati.append(valore)

        file = request.files.get("filepdf")
        filepdf_name = ""
        pdf_caricato = False
        if file and allowed_file(file.filename):
            base_name = secure_filename(file.filename.rsplit(".", 1)[0])
            filepdf_name = f"{base_name}.pdf"
            save_path = os.path.join(PDF_FOLDER, filepdf_name)
            file.save(save_path)
            pdf_caricato = True

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            INSERT INTO norme (anno, numero, tipologia, argomento, oggetto,
                               descrizione, stato, note, fonte, filepdf)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (*dati, filepdf_name))
        conn.commit()
        new_id = c.lastrowid
        conn.close()

        if os.path.exists(EXCEL_FILE):
            try:
                df = pd.read_excel(EXCEL_FILE)
                nuova_riga = pd.DataFrame([{
                    "Anno": dati[0], "Numero": dati[1], "Tipologia": dati[2],
                    "Argomento": dati[3], "Oggetto": dati[4], "Descrizione": dati[5],
                    "Stato": dati[6], "Note": dati[7], "Fonte": dati[8],
                    "FilePDF": filepdf_name
                }])
                df = pd.concat([df, nuova_riga], ignore_index=True)
                df.to_excel(EXCEL_FILE, index=False)
            except Exception:
                traceback.print_exc()

        log_event("insert", norma_id=new_id, details={
            "anno": dati[0], "numero": dati[1], "tipologia": dati[2],
            "argomento": dati[3], "has_pdf": pdf_caricato, "filepdf": filepdf_name
        })

        # 🔓 Rilascio lock dopo l'inserimento del nuovo atto
        try:
            release_lock()
        except Exception:
            traceback.print_exc()

        if pdf_caricato:
            flash("📎 PDF caricato correttamente!", "success")
        flash("✅ Nuovo atto inserito con successo!", "success")
        return redirect(url_for("ricerca"))

    return render_template("inserisci_norma.html", tipologie=tipologie, categorie=argomenti, fonti=fonti)

# ==========================================================
# 📎 PDF (con Audit)
# ==========================================================
@app.route("/carica_pdf/<int:norma_id>", methods=["GET", "POST"])
@require_write_lock
def carica_pdf(norma_id):
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    norma = c.execute("SELECT * FROM norme WHERE id=?", (norma_id,)).fetchone()
    conn.close()

    if request.method == "POST":
        file = request.files.get("filepdf")
        if not file or not allowed_file(file.filename):
            flash("⚠️ Caricare un file PDF valido.", "error")
            return redirect(url_for("carica_pdf", norma_id=norma_id))
        try:
            base_name = secure_filename(file.filename.rsplit(".", 1)[0])
            final_name = f"{base_name}.pdf"
            save_path = os.path.join(PDF_FOLDER, final_name)
            file.save(save_path)

            # aggiorna DB
            conn = sqlite3.connect(DB_FILE)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            old_row = c.execute("SELECT filepdf FROM norme WHERE id=?", (norma_id,)).fetchone()
            old_name = old_row["filepdf"] if old_row else ""
            c.execute("UPDATE norme SET filepdf=? WHERE id=?", (final_name, norma_id))
            conn.commit()
            conn.close()

            # audit pdf_upload/pdf_replace
            action = "pdf_replace" if (old_name or "") else "pdf_upload"
            log_event(action, norma_id=norma_id, details={"old": old_name or "", "new": final_name})

            # aggiorna Excel (best effort)
            try:
                if os.path.exists(EXCEL_FILE) and norma:
                    df = pd.read_excel(EXCEL_FILE)
                    mask = (df["Anno"].astype(str) == str(norma["anno"])) & \
                           (df["Numero"].astype(str) == str(norma["numero"]))
                    if "FilePDF" in df.columns and mask.any():
                        df.loc[mask, "FilePDF"] = final_name
                        df.to_excel(EXCEL_FILE, index=False)
            except Exception:
                traceback.print_exc()

            backup_excel()
            backup_pdf()

            # 🔓 Rilascio lock dopo il caricamento del PDF
            try:
                release_lock()
            except Exception:
                traceback.print_exc()

            flash("📎 PDF caricato correttamente!", "success")
        except Exception as e:
            traceback.print_exc()
            flash(f"❌ Errore durante il caricamento PDF: {e}", "error")
        return redirect(url_for("dettaglio", norma_id=norma_id))

    return render_template("carica_pdf.html", norma=norma)

@app.route("/pdf/<filename>")
def apri_pdf(filename):
    pdf_folder = PDF_FOLDER
    if not filename.lower().endswith(".pdf"):
        filename += ".pdf"
    path = os.path.join(pdf_folder, filename)
    if not os.path.exists(path):
        flash(f"⚠️ Il file PDF \"{filename}\" non è stato trovato nella cartella PDF condivisa.", "error")
        return redirect(url_for("ricerca"))
    return send_from_directory(pdf_folder, filename)

# ==========================================================
# 👀 AUDIT VIEW (solo admin) — endpoint + alias robusti
# ==========================================================
def _render_audit_table(rows):
    # Fallback HTML se manca il template audit.html
    html = ["<h2>Audit log (ultimi 300 eventi)</h2>",
            "<table border='1' cellpadding='6' cellspacing='0'>",
            "<tr><th>#</th><th>Quando</th><th>Azione</th><th>ID Atto</th>"
            "<th>Utente</th><th>IP</th><th>User-Agent</th><th>Dettagli</th></tr>"]
    for e in rows:
        html.append(
            f"<tr><td>{e['id']}</td><td>{e['ts']}</td><td>{e['action']}</td>"
            f"<td>{e['norma_id'] or ''}</td><td>{e['actor']}</td><td>{e['ip']}</td>"
            f"<td style='font-size:12px;color:#666'>{(e['user_agent'] or '')}</td>"
            f"<td><pre style='white-space:pre-wrap'>{e['details'] or ''}</pre></td></tr>"
        )
    html.append("</table><p><a href='/admin_dashboard'>⬅︎ Torna alla dashboard</a></p>")
    return "\n".join(html)

def audit_view_handler():
    if not session.get("admin"):
        flash("⚠️ Accesso non autorizzato.", "warning")
        return redirect(url_for("admin_login"))

    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT * FROM audit ORDER BY id DESC LIMIT 300").fetchall()
    conn.close()
    try:
        return render_template("audit.html", events=rows)
    except TemplateNotFound:
        return _render_audit_table(rows)

# Registra SEMPRE entrambi gli endpoint (se non esistono già).
def _ensure_audit_routes():
    existing = {r.endpoint for r in app.url_map.iter_rules()}
    if "audit_view" not in existing:
        app.add_url_rule("/audit", view_func=audit_view_handler, endpoint="audit_view", methods=["GET"])
    if "admin_audit" not in existing:
        app.add_url_rule("/admin/audit", view_func=audit_view_handler, endpoint="admin_audit", methods=["GET"])

_ensure_audit_routes()

# (facoltativo) helper per il template se vuoi usare url_for dinamico
@app.context_processor
def inject_endpoint_exists():
    # Sola lettura = siamo online ma questa istanza NON ha il lock di scrittura.
    readonly = is_online() and (not HAS_WRITE_LOCK)
    return {
        "endpoint_exists": lambda name: name in app.view_functions,
        "read_only": readonly,
        "can_write": not readonly,
    }

# ==========================================================
# 🔁 SECONDA /boot-ready (rinominata per non confliggere)
# ==========================================================
@app.route("/boot-ready-redirect.js")
def boot_ready_redirect_js():
    js = """
    if (window.location.pathname !== "/") {
        window.location.replace("/");
    } else {
        console.log("redirect evitato: già in /");
    }
    """
    resp = make_response(js, 200)
    resp.headers["Content-Type"] = "application/javascript"
    resp.headers["Cache-Control"] = "no-store"
    return resp

# ==========================================================
# AVVIO
# ==========================================================

@app.post("/kill-python")
def kill_python():
    import os
    os._exit(0)

if __name__ == "__main__":
    print("🔎 Verifica rete…")

    if is_online():
        print("🌐 Sei ONLINE → provo a sincronizzare da rete.")
        sync_from_network()
    else:
        print("📴 OFFLINE → uso solo database locale.")

    crea_database()
    ensure_audit_table()

    app.run(debug=True, port=5001, use_reloader=False)

    # Alla chiusura prova a sincronizzare locale → rete (solo se online e con lock)
    sync_to_network()